import os
import sys
import csv
import glob
import json
import openpyxl
from ui import Mainui
from PyQt5.QtCore import QTimer
from easydict import EasyDict as edict
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import QApplication, QFileDialog, QLabel, QMessageBox, QAbstractItemView


class Fun(Mainui):
    def __init__(self):
        super().__init__()
        self.setupUi()
        self.clean_result_uuid_list = []
        self.send_rcc_uuid_list = []
        self.base_img_label = {"large_garbage": "垃圾", "garbage": "垃圾", "sewage": "脏污"}
        self.same_list = []
        self.csv_data = {}
        self.show()

        self.action_Log.triggered.connect(self.get_log_file_path)
        self.action.triggered.connect(self.get_save_result_path)
        self.action_CSV.triggered.connect(self.load_csv_file)
        self.pushButton.clicked.connect(self.save_result_file)
    
    def load_csv_file(self):
        if self.input_img.text() == "":
            reply = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("请先选择图片所在目录 !"), QMessageBox.NoButton, self)
            yr_btn = reply.addButton(self.tr("选择图片目录"), QMessageBox.YesRole)
            reply.addButton(self.tr("取消"), QMessageBox.NoRole)
            reply.exec_()
            if reply.clickedButton() == yr_btn:
                self.load_img_files()

        file = QFileDialog.getOpenFileName(self, "选择 info.csv 文件", filter="Text Files (*.csv)")
        if file is not None:
            self.input_csv.setText(file[0])
            QTimer.singleShot(1000, self.read_csv_img)
        self.listen_path()

    def load_img_files(self):
        path = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
        if path is not None:
            self.input_img.setText(path)

        self.listen_path()
    
    def read_csv_img(self):
        data = {}
        try:
            with open(self.input_csv.text(), 'r') as f:
                reader = list(csv.reader(f))[1:]
            img_list = glob.glob(f'{self.input_img.text()}\\**')
            imglist = [item.split('.')[0] for dirl in img_list for item in os.listdir(dirl)]

            for row in reader:
                image_name = row[1]
                class_name = row[2]
                uuid = row[4]

                if image_name in imglist:
                    data[uuid] = {"img_name": image_name, "class_image": class_name}
            self.csv_data = data
            if self.input_log.text() == "":
                reply = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("info.csv 已加载完成, 请加载log文件 !"), QMessageBox.NoButton, self)
                yr_btn = reply.addButton(self.tr("请选择 log 所在目录"), QMessageBox.YesRole)
                reply.addButton(self.tr("取消"), QMessageBox.NoRole)
                reply.exec_()
                if reply.clickedButton() == yr_btn:
                    self.load_log_files()
            else:
                self.load_log_files()
        except:
            QMessageBox.warning(self, self.tr("警告"), self.tr("是否加载了正确的 info.csv 文件 !"), QMessageBox.Ok)

    def load_log_files(self):
        if self.input_csv.text() == "":
            reply = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("请先 加载 info.csv 文件 !"), QMessageBox.NoButton, self)
            yr_btn = reply.addButton(self.tr("选择 info.csv 文件"), QMessageBox.YesRole)
            reply.addButton(self.tr("取消"), QMessageBox.NoRole)
            reply.exec_()
            if reply.clickedButton() == yr_btn:
                self.load_csv_file()
        else:
            logfiles = glob.glob(f'{self.input_log.text()}/gs_console*.log')
            
            if len(logfiles) == 0:
                reply = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("该目录下没有 console 日志 , 请重选目录"), QMessageBox.NoButton, self)
                yr_btn = reply.addButton(self.tr("重选"), QMessageBox.YesRole)
                reply.addButton(self.tr("取消"), QMessageBox.NoRole)
                reply.exec_()
                if reply.clickedButton() == yr_btn:
                    self.get_log_file_path()
                else:
                    self.input_log.setText("")
            else:
                for item in logfiles:
                    self.show_message(f"正在加载 {item} 文件")
                    with open(item, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    for l in lines:
                        if "[garbage_upload] send rcc" in l:
                            tmp = edict(json.loads(l.split(" ")[-1].replace("\n", "")))
                            t_l = [tmp.id, tmp.category]
                            if t_l not in self.send_rcc_uuid_list:
                                self.send_rcc_uuid_list.append(t_l)
                        
                        if "clean_result(0)" in l:
                            tmp = l.split(" ")[-2].replace("dirty_id(", "").replace(")", "")
                            if tmp not in self.clean_result_uuid_list:
                                self.clean_result_uuid_list.append(tmp)
                for item in self.send_rcc_uuid_list:
                    if item[0] in self.clean_result_uuid_list:
                        self.same_list.append(item[0])
                QTimer.singleShot(1000, lambda:self.show_message("加载 log 文件完成"))
                
                QTimer.singleShot(2000, self.show_label)

                QTimer.singleShot(3000, self.create_table)

            self.listen_path()
    
    def create_table(self):
        length_clean_result = len(self.clean_result_uuid_list)
        model = QStandardItemModel(self.tableView)
        model.setHorizontalHeaderLabels(['UUID', '标签', '中文标签', '图片名'])
        model.setColumnCount(4)
        model.setRowCount(length_clean_result)
        for i in range(length_clean_result):
            item1 = QStandardItem(self.clean_result_uuid_list[i])
            if self.clean_result_uuid_list[i] in self.csv_data:
                tmp_dict = edict(self.csv_data[self.clean_result_uuid_list[i]])
                self.clean_result_uuid_list[i] = [self.clean_result_uuid_list[i], tmp_dict.class_name, self.base_img_label[tmp_dict.class_name], tmp_dict.img_name]
                item2 = QStandardItem(tmp_dict.class_name)
                item3 = QStandardItem(self.base_img_label[tmp_dict.class_name])
                item4 = QStandardItem(tmp_dict.img_name)
            else:
                self.clean_result_uuid_list[i] = (self.clean_result_uuid_list[i], "", "", "")
                item2 = QStandardItem("")
                item3 = QStandardItem("")
                item4 = QStandardItem("")
            model.setItem(i, 0, item1)
            model.setItem(i, 1, item2)
            model.setItem(i, 2, item3)
            model.setItem(i, 3, item4)
        self.tableView.setModel(model)
        self.tableView.setColumnWidth(0, 300)
        self.tableView.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableView.show()

        # ===========================================================

        length_send_rcc = len(self.send_rcc_uuid_list)
        model_2 = QStandardItemModel(self.tableView_2)
        model_2.setColumnCount(4)
        model_2.setRowCount(length_send_rcc)
        model_2.setHorizontalHeaderLabels(['UUID', '标签', '中文标签', '图片名'])
        
        for i in range(length_send_rcc):
            tmp = self.send_rcc_uuid_list[i]
            tmp.append(self.base_img_label[tmp[-1]])

            item1 = QStandardItem(tmp[0])
            item2 = QStandardItem(tmp[1])
            item3 = QStandardItem(tmp[2])
            if tmp[0] in self.csv_data:
                tmp.append(self.csv_data[tmp[0]]['img_name'])
                item4 = QStandardItem(self.csv_data[tmp[0]]['img_name'])
            else:
                tmp.append("")
                item4 = QStandardItem("")

            self.send_rcc_uuid_list[i] = tmp

            model_2.setItem(i, 0, item1)
            model_2.setItem(i, 1, item2)
            model_2.setItem(i, 2, item3)
            model_2.setItem(i, 3, item4)
        self.tableView_2.setModel(model_2)
        self.tableView_2.setColumnWidth(0, 300)
        self.tableView_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableView_2.show()
        # ===================================================================
        length_same = len(self.same_list)
        model_3 = QStandardItemModel(self.tableView_3)
        model_3.setColumnCount(4)
        model_3.setRowCount(length_same)
        model_3.setHorizontalHeaderLabels(['UUID', '标签', '中文标签', '图片名'])
        
        for i in range(length_same):
            item1 = QStandardItem(self.same_list[i])
            if self.clean_result_uuid_list[i] in self.csv_data:
                dict_tmp = edict(self.csv_data[self.same_list[i]])
                self.same_list[i] = [self.same_list[i], dict_tmp.class_name, self.base_img_label[dict_tmp.class_name], self.base_img_label[dict_tmp.img_name]]
                item2 = QStandardItem(dict_tmp.class_name)
                item3 = QStandardItem(self.base_img_label[dict_tmp.class_name])
                item4 = QStandardItem(self.base_img_label[dict_tmp.img_name])
            else:
                self.same_list[i] = [self.same_list[i], "", "", ""], 
                item2 = QStandardItem("")
                item3 = QStandardItem("")
                item4 = QStandardItem("")

            model_3.setItem(i, 0, item1)
            model_3.setItem(i, 1, item2)
            model_3.setItem(i, 2, item3)
            model_3.setItem(i, 3, item4)
    
        self.tableView_3.setModel(model_3)
        self.tableView_3.setColumnWidth(0, 300)
        self.tableView_3.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableView_3.show()

        self.csv_data = {}

    def show_message(self, msg):
        self.statusBar().showMessage(msg)
    
    def show_label(self):
        self.label_status = QLabel(f"clean_result:{len(self.clean_result_uuid_list)} 个 || send_rcc: {len(self.send_rcc_uuid_list)} 个 || same: {len(self.same_list)} 个")
        self.statusBar().addPermanentWidget(self.label_status)

    def get_log_file_path(self):
        path = QFileDialog.getExistingDirectory(self, "选择 Log 所在文件夹")
        if path is not None:
            self.input_log.setText(path)

            QTimer.singleShot(1000, self.load_log_files)

    def get_save_result_path(self):
        path = QFileDialog.getExistingDirectory(self, "选择保存文件夹")
        if path is not None:
            self.input_result.setText(path)
    
    def listen_path(self):
        log_dir = self.input_log.text()
        csv_dir = self.input_csv.text()
        img_dir = self.input_img.text()
        if log_dir != "" and csv_dir != "" and img_dir != "":
            self.pushButton.setDisabled(False)
        else:
            self.pushButton.setDisabled(True)

    def save_result_file(self):
        if self.input_result.text() == "":
            reply = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("请选择结果保存目录 !"), QMessageBox.NoButton, self)
            yr_btn = reply.addButton(self.tr("选择保存目录"), QMessageBox.YesRole)
            reply.addButton(self.tr("取消"), QMessageBox.NoRole)
            reply.exec_()
            if reply.clickedButton() == yr_btn:
                self.get_save_result_path()
            else:
                self.input_result.setText("")
        else:
            workbook = openpyxl.Workbook()

            sheet = workbook.active
            sheet.title = "send rcc"

            sheet['A1'] = "UUID"
            sheet['B1'] = "标签"
            sheet['C1'] = "中文标签"
            sheet['D1'] = "图片名"

            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['D'].width = 15

            for it in self.send_rcc_uuid_list:
                sheet.append(it)

            sheet2 = workbook.create_sheet("clean result")
            sheet2['A1'] = "UUID"
            sheet2['B1'] = "标签"
            sheet2['C1'] = "中文标签"
            sheet2['D1'] = "图片名"

            sheet2.column_dimensions['A'].width = 40
            sheet2.column_dimensions['B'].width = 15
            sheet2.column_dimensions['D'].width = 15

            for it in self.clean_result_uuid_list:
                sheet2.append(it)
            
            sheet3 = workbook.create_sheet("same")

            sheet3['A1'] = "UUID"
            sheet3['B1'] = "标签"
            sheet3['C1'] = "中文标签"
            sheet3['D1'] = "图片名"

            sheet3.column_dimensions['A'].width = 40
            sheet3.column_dimensions['B'].width = 15
            sheet3.column_dimensions['D'].width = 15

            for it in self.same_list:
                sheet3.append(it)

            try:
                workbook.save(f"{self.input_result.text()}/show.xlsx")
                reply = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("已保存到该目录下, 名为 show.xlsx !"), QMessageBox.NoButton, self)
                yr_btn = reply.addButton(self.tr("确定"), QMessageBox.YesRole)
                reply.exec_()
            except PermissionError as pe:
                reply = QMessageBox(QMessageBox.Question, self.tr("警告"), self.tr(str(pe)), QMessageBox.NoButton, self)
                yr_btn = reply.addButton(self.tr("确定"), QMessageBox.YesRole)
                reply.exec_()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    fun = Fun()
    sys.exit(app.exec_())